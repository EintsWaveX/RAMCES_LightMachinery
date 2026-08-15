"""
READS and CLEANS `modules/KATALOG SFM.xlsx`. Writes nothing.

Same contract as `seed_aset_real.py`, which owns the asset workbook: this module
opens a spreadsheet, cleans it, and hands back plain tuples. It imports no
writer and touches no database, so `py -3.10 seed_katalog_sfm.py` is a safe dry
run against the client's file.

WHY THIS EXISTS
---------------
Until rev0.5.2 this workbook was never opened by any code. All 373 of its rows
reached the database through a hand transcription in `seed_katalog.py`, and that
transcription had already drifted from the file it claimed to mirror:

  * 14 MEKANIK resorts (ME1.1, ME1.2, ME2..ME9, MEI..MEIV) were missing
    entirely — the sheet has 254 UPTs, the database had 240 of them;
  * DHL and DHS carry `Alat Ukur = TRUE` in the sheet and `False` in the code;
  * BY1 is "BALAIYASA CIREBON PRUJAKAN" in the sheet and
    "BALAIYASA CIREBONPRUNJAKAN" in the code.

None of those were noticed for a release, because nothing compared the two. The
transcription is now gone and this is the only path in.

THE THREE SHEETS
----------------
  KATALOG ALAT KERJA   A:D  103 tool types + a hidden H/I side table
  KATALOG LOKASI       A:C   16 parent regions
  KATALOG UPT          A:D  254 resorts

WHAT STILL LIVES IN seed_katalog.py
-----------------------------------
Only what is genuinely NOT in the workbook: the three minted codes for rows the
client left without one, the BKC row the katalog itself marks "Tidak Ditemukan",
the EXTRA_ALAT_REMAP table, and the Model/Type data transcribed from
`Rekap Spek RAMCES.docx`.
"""

import os
import re
from collections import namedtuple

try:
    import openpyxl
except ImportError:  # pragma: no cover - openpyxl is installed on py -3.10
    openpyxl = None


WORKBOOK = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "modules", "KATALOG SFM.xlsx"
)

SHEET_ALAT = "KATALOG ALAT KERJA"
SHEET_LOKASI = "KATALOG LOKASI"
SHEET_UPT = "KATALOG UPT"

# Expected shapes. These are assertions, not documentation: a client drop that
# changes them must fail the seed loudly rather than import a different fleet.
N_ALAT = 103
N_LOKASI = 16
N_UPT = 254

# The H/I side table has NO header anywhere in the sheet. It is two runs of
# rows separated by blank ones, and the two runs correspond to the two
# subfolders of `modules/Spesifikasi Lengkap Alat Kerja/`. The mapping is
# therefore POSITIONAL, which is why read_kelompok() refuses to guess when the
# block count changes.
KELOMPOK_BLOK = ("FASILITAS", "JEMBATAN")


KatalogSFM = namedtuple("KatalogSFM", "alat lokasi upt kelompok minted laporan")
#   alat     [(kode, nama, alat_ukur, perlu_kalibrasi, kelompok)]
#   lokasi   [(kode, nama, tipe)]
#   upt      [(kode, nama, parent_kode, parent_nama, unit)]
#   kelompok {kode: "FASILITAS" | "JEMBATAN" | "FASILITAS,JEMBATAN"}
#   minted   {nama: kode}  the blank-Kode rows and what they were given
#   laporan  {counter: value}, printed by the __main__ dry run


# ══════════════════════════════════════════════════════════════════════
# Cleaning
# ══════════════════════════════════════════════════════════════════════
#
# Two hazards in this particular file, both silent:
#
#   * `Induk Lokasi` reads "DAOP 1 JAKARTA\xa0(D1)" — the separator before the
#     bracket is U+00A0, a NON-BREAKING space. `.split(" ")` yields "DAOP".
#     Python's `\s` DOES match U+00A0, so normalising with re is safe; splitting
#     on a literal space is not.
#   * exactly one code in 254 rows carries a trailing space ("JB1.3 "). Cleaning
#     has to happen BEFORE the duplicate check or it reads as a 255th resort.


def _clean_code(v):
    """A code with every kind of whitespace removed, upper-cased."""
    if v is None:
        return ""
    return re.sub(r"\s+", "", str(v)).upper()


def _clean_text(v):
    """Display text with runs of whitespace (including U+00A0) collapsed."""
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\xa0", " ")).strip()


def _as_bool(v):
    """The sheet's boolean columns, which arrive as bool, 'TRUE'/'YA', or blank."""
    if isinstance(v, bool):
        return v
    s = _clean_code(v)
    return s in {"TRUE", "YA", "Y", "1", "ADA", "V"}


# `NAMA\xa0(KODE)` — anchored, because a name may itself contain brackets.
_RE_INDUK = re.compile(r"^(?P<nama>.+?)\s*\((?P<kode>[A-Z0-9]+)\)$", re.IGNORECASE)


def _parse_induk(cell):
    """'DAOP 1 JAKARTA\\xa0(D1)' -> ('D1', 'DAOP 1 JAKARTA'). Raises on anything else."""
    teks = _clean_text(cell)
    m = _RE_INDUK.match(teks)
    if not m:
        raise ValueError(f"kolom 'Induk Lokasi' tidak bisa dibaca: {cell!r}")
    return m.group("kode").upper(), m.group("nama").strip()


def _require_workbook(path):
    if openpyxl is None:
        raise RuntimeError(
            "openpyxl tidak terpasang. Jalankan dengan `py -3.10`, bukan `python`."
        )
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"{path} tidak ditemukan.\n"
            "modules/ adalah drop data milik klien dan tidak ikut di-commit. "
            "Salin folder itu ke root proyek sebelum menjalankan seed."
        )


# ══════════════════════════════════════════════════════════════════════
# The three sheets
# ══════════════════════════════════════════════════════════════════════


def read_alat(ws, kode_tanpa_kode, kelompok):
    """A:D -> [(kode, nama, alat_ukur, perlu_kalibrasi, kelompok)] + what was minted.

    Three rows in the sheet have a name and no code (GAUGE CALIBRATOR, TRACKER,
    WATER STEAM). Their codes come from `kode_tanpa_kode` in seed_katalog.py and
    a MISS IS A HARD ERROR — never invent one from the name, or a fourth blank
    row in a future drop silently becomes "GAU" and starts collecting assets.
    """
    rows, minted, seen = [], {}, {}
    for r in range(2, ws.max_row + 1):
        kode = _clean_code(ws.cell(r, 1).value)
        nama = _clean_text(ws.cell(r, 2).value)
        if not kode and not nama:
            continue

        if not kode:
            kode = kode_tanpa_kode.get(nama.upper())
            if not kode:
                raise ValueError(
                    f"{SHEET_ALAT} baris {r}: tidak ada Kode dan tidak ada kode "
                    f"cadangan untuk {nama!r}. Tambahkan ke KODE_TANPA_KODE di "
                    f"seed_katalog.py."
                )
            minted[nama] = kode

        if kode in seen:
            raise ValueError(
                f"{SHEET_ALAT}: kode {kode!r} muncul dua kali "
                f"(baris {seen[kode]} dan {r})"
            )
        seen[kode] = r

        rows.append(
            (
                kode,
                nama,
                _as_bool(ws.cell(r, 3).value),
                _as_bool(ws.cell(r, 4).value),
                kelompok.get(kode),
            )
        )
    return rows, minted


def read_kelompok(ws, blok_labels=KELOMPOK_BLOK):
    """The hidden H/I side table -> {kode: 'FASILITAS' | 'JEMBATAN' | both}.

    The sheet names neither group. They are two runs of rows in column H
    (H34:H44 and H47:H59 in the current drop) separated by blank rows, and they
    correspond, in order, to the two subfolders of
    `modules/Spesifikasi Lengkap Alat Kerja/`.

    Because that mapping is positional, an unexpected number of blocks must
    RAISE. Silently tagging 24 tool types with the wrong kelompok would change
    which spektek document each one offers, and nothing downstream would notice.
    """
    blocks, current = [], []
    for r in range(2, ws.max_row + 1):
        kode = _clean_code(ws.cell(r, 8).value)
        if kode:
            current.append((r, kode))
        elif current:
            blocks.append(current)
            current = []
    if current:
        blocks.append(current)

    if len(blocks) != len(blok_labels):
        rentang = [(b[0][0], b[-1][0]) for b in blocks]
        raise ValueError(
            f"{SHEET_ALAT} kolom H: diharapkan {len(blok_labels)} blok "
            f"({', '.join(blok_labels)}), ditemukan {len(blocks)} pada baris {rentang}. "
            f"Pembagian blok bersifat POSISIONAL — perbarui KELOMPOK_BLOK bila "
            f"klien mengubah susunannya."
        )

    out = {}
    for label, block in zip(blok_labels, blocks):
        for _r, kode in block:
            if kode in out:
                if label not in out[kode].split(","):
                    out[kode] = ",".join(sorted(set(out[kode].split(",") + [label])))
            else:
                out[kode] = label
    return out


def read_lokasi(ws):
    """A:C -> [(kode, nama, tipe)] for the 16 parent regions."""
    rows, seen = [], set()
    for r in range(2, ws.max_row + 1):
        kode = _clean_code(ws.cell(r, 1).value)
        nama = _clean_text(ws.cell(r, 2).value)
        tipe = _clean_text(ws.cell(r, 3).value).upper()
        if not kode:
            continue
        if kode in seen:
            raise ValueError(f"{SHEET_LOKASI}: kode {kode!r} muncul dua kali")
        seen.add(kode)
        rows.append((kode, nama, tipe))
    return rows


def read_upt(ws):
    """A:D -> [(kode, nama, parent_kode, parent_nama, unit)] for the 254 resorts.

    `unit` is the sheet's own UNIT column (JALAN REL / JEMBATAN / MEKANIK). It
    maps 1:1 onto the code prefix (JR / JB / ME) across all 254 rows, which is
    what makes `peruntukan_dari_lokasi()` below a derivation rather than a guess.
    """
    rows, seen = [], {}
    for r in range(2, ws.max_row + 1):
        unit = _clean_text(ws.cell(r, 1).value).upper()
        kode = _clean_code(ws.cell(r, 2).value)  # "JB1.3 " -> "JB1.3"
        nama = _clean_text(ws.cell(r, 3).value)
        if not kode:
            continue
        if kode in seen:
            raise ValueError(
                f"{SHEET_UPT}: kode {kode!r} muncul dua kali "
                f"(baris {seen[kode]} dan {r})"
            )
        seen[kode] = r
        parent_kode, parent_nama = _parse_induk(ws.cell(r, 4).value)
        rows.append((kode, nama, parent_kode, parent_nama, unit))
    return rows


# ══════════════════════════════════════════════════════════════════════
# The peruntukan rule
# ══════════════════════════════════════════════════════════════════════
#
# THE UPT CODE PREFIX IS THE PERUNTUKAN. Verified 1:1 against the sheet's own
# UNIT column across all 254 rows:
#
#     JR*  ->  JALAN REL  (A)   202 rows
#     JB*  ->  JEMBATAN   (B)    38 rows
#     ME*  ->  MEKANIK    (C)    14 rows
#
# This matters because the ASSET workbook's own `Unit Peruntukan` column is not
# reliable: every one of the 245 rows on `Data Aset JB` carries "A" even though
# they are bridge assets on JB resorts. `peruntukan` is baked into the composite
# primary key, so importing that column verbatim mints 245 permanently wrong
# keys.
#
# BY* IS NOT A UNIT. Balaiyasa is a workshop; an asset parked at BY3A is
# visiting for repair and keeps the peruntukan it came with. Reclassifying it
# by location would contradict the rule enforced in five other places that a
# Balaiyasa is never a reporting region. For BY* — and for a bare parent code
# like D1 — this returns None and the caller keeps the workbook's own value.

_PREFIX_PERUNTUKAN = (
    ("JR", "JALAN REL"),
    ("JB", "JEMBATAN"),
    ("ME", "MEKANIK"),
)


def peruntukan_dari_lokasi(id_lokasi):
    """UPT code -> peruntukan, or None when the code does not encode one."""
    kode = _clean_code(id_lokasi)
    for prefix, peruntukan in _PREFIX_PERUNTUKAN:
        if kode.startswith(prefix):
            return peruntukan
    return None


# ══════════════════════════════════════════════════════════════════════
# Entry point
# ══════════════════════════════════════════════════════════════════════

_cache = {}


def read_katalog(path=WORKBOOK, kode_tanpa_kode=None, force=False):
    """Read the whole workbook once and memoise it.

    Memoised because `seeds/katalog.py`, `seeds/verify.py` and `seed_aset_real`
    all want it within one process and the file is ~95 KB of XML.
    """
    if not force and path in _cache:
        return _cache[path]

    if kode_tanpa_kode is None:
        # Imported lazily: seed_katalog.py imports THIS module for the reader,
        # so a module-level import here would be a cycle.
        from seed_katalog import KODE_TANPA_KODE

        kode_tanpa_kode = KODE_TANPA_KODE

    _require_workbook(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        for nama_sheet in (SHEET_ALAT, SHEET_LOKASI, SHEET_UPT):
            if nama_sheet not in wb.sheetnames:
                raise ValueError(
                    f"sheet {nama_sheet!r} tidak ada di {os.path.basename(path)}. "
                    f"Sheet yang ada: {wb.sheetnames}"
                )

        ws_alat = wb[SHEET_ALAT]
        kelompok = read_kelompok(ws_alat)
        alat, minted = read_alat(ws_alat, {k.upper(): v for k, v in kode_tanpa_kode.items()}, kelompok)
        lokasi = read_lokasi(wb[SHEET_LOKASI])
        upt = read_upt(wb[SHEET_UPT])
    finally:
        wb.close()

    # Shape assertions. A client drop that changes these is a real event and
    # must stop the seed, not quietly import a different fleet.
    for label, got, want in (
        ("alat kerja", len(alat), N_ALAT),
        ("lokasi", len(lokasi), N_LOKASI),
        ("UPT", len(upt), N_UPT),
    ):
        if got != want:
            raise ValueError(
                f"{os.path.basename(path)}: {label} berjumlah {got}, diharapkan {want}. "
                f"Bila klien memang mengubah katalog, perbarui N_* di "
                f"seed_katalog_sfm.py dan jalankan `py -3.10 seed_katalog_sfm.py` "
                f"untuk melihat selisihnya."
            )

    unit_counts = {}
    for _k, _n, _pk, _pn, unit in upt:
        unit_counts[unit] = unit_counts.get(unit, 0) + 1

    hasil = KatalogSFM(
        alat=alat,
        lokasi=lokasi,
        upt=upt,
        kelompok=kelompok,
        minted=minted,
        laporan={
            "alat": len(alat),
            "lokasi": len(lokasi),
            "upt": len(upt),
            "unit": unit_counts,
            "kelompok": len(kelompok),
            "alat_ukur": sum(1 for r in alat if r[2]),
            "perlu_kalibrasi": sum(1 for r in alat if r[3]),
            "minted": dict(minted),
        },
    )
    _cache[path] = hasil
    return hasil


def verify_induk(katalog=None):
    """Every UPT's parsed `Induk Lokasi` must exist in KATALOG LOKASI, by code AND name.

    Catches a rename applied to one sheet and not the other — which is exactly
    how the BY1 spelling drifted between the workbook and the transcription.
    """
    k = katalog or read_katalog()
    by_code = {kode: nama for kode, nama, _tipe in k.lokasi}
    bad = []
    for kode, _nama, parent_kode, parent_nama, _unit in k.upt:
        if parent_kode not in by_code:
            bad.append((kode, parent_kode, "kode induk tidak ada di KATALOG LOKASI"))
        elif by_code[parent_kode].upper() != parent_nama.upper():
            bad.append(
                (kode, parent_kode,
                 f"nama induk berbeda: {parent_nama!r} vs {by_code[parent_kode]!r}")
            )
    return bad


if __name__ == "__main__":
    # Dry run: read and validate the workbook without touching the database.
    k = read_katalog()
    lap = k.laporan
    print(f"alat kerja      : {lap['alat']}")
    print(f"  alat ukur     : {lap['alat_ukur']}")
    print(f"  perlu kalibrasi: {lap['perlu_kalibrasi']}")
    print(f"  kode diciptakan: {lap['minted'] or '—'}")
    print(f"kelompok (H/I)  : {lap['kelompok']} kode")
    print(f"lokasi induk    : {lap['lokasi']}")
    print(f"UPT             : {lap['upt']}")
    for unit, n in sorted(lap["unit"].items()):
        print(f"  {unit:<12}: {n}")

    bad = verify_induk(k)
    print(f"induk tidak cocok: {len(bad)}")
    for kode, parent, why in bad[:10]:
        print(f"   {kode} -> {parent}: {why}")

    # The peruntukan derivation, as a table rather than a claim.
    print("\nperuntukan dari prefix kode UPT:")
    cek = {}
    for kode, _n, _pk, _pn, unit in k.upt:
        cek.setdefault((peruntukan_dari_lokasi(kode), unit), 0)
        cek[(peruntukan_dari_lokasi(kode), unit)] += 1
    for (derived, unit), n in sorted(cek.items(), key=lambda x: str(x[0])):
        tanda = "OK " if derived == unit else "BEDA"
        print(f"   {tanda} prefix->{derived!s:<10} UNIT={unit:<10} {n} baris")
