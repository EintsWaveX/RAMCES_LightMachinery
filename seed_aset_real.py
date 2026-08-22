"""
Import the real fleet from `modules/Template_Import_Aset_RAMCES.xlsx`.

Replaces the randomly generated demo assets. 1,121 rows across four sheets,
and every one of them needed cleaning, the four sheets were maintained
separately and disagree about almost everything:

  * Dates arrive as real `datetime`, as "21/4/26", as "24/11/2025", as
    "2026-12-31", and, 160 times, as the literal string "PROSES".
  * `ID Lokasi` is filled on two sheets and blank on the other two, which only
    carry `Parent Lokasi`.
  * `Parent Lokasi` uses DI/DII/DIII/DIV on one sheet where the schema wants
    VI/VII/VIII/VIV, and "KAC" for Balaiyasa Kiaracondong (BY3).
  * UPT codes carry stray whitespace: "JB1.3 ", "JB 1.4".

Everything that cannot be resolved is SKIPPED AND COUNTED, never invented, the
one exception is documented at PROSES_FALLBACK below.
"""

import os
import re
from datetime import date, datetime

import models
from seed_katalog import ALL_KODE
from seed_katalog_sfm import peruntukan_dari_lokasi

WORKBOOK = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "modules",
    "Template_Import_Aset_RAMCES.xlsx",
)

SHEETS = ["Data Aset", "Data Aset 2026", "Data Aset 2025", "Data Aset JB"]

# A sheet that is itself a register of one unit. `Data Aset JB` is the client's
# jembatan list, 245 rows, every one of them a bridge asset, even though its
# `Unit Peruntukan` column says "A" on all of them.
#
# This is the tie-breaker for the five rows that are parked at BY3A for repair:
# their location is a workshop, so the code prefix cannot say what they are, and
# the Unit column is the same "A" that is wrong for the whole sheet. Without
# this they would be the only five bridge tools in the fleet with JALAN REL
# baked into their primary key.
SHEET_PERUNTUKAN = {"Data Aset JB": "JEMBATAN"}

# Column indexes, 0-based. Identical across all four sheets; only the header
# TEXT differs (the JB sheet has an extra space in every caption).
COL_KODE, COL_URUT, COL_PENGADAAN = 1, 2, 3
COL_TANGGAL, COL_UNIT, COL_IDLOK, COL_PARENT, COL_MODEL = 4, 5, 6, 7, 8

# The 2026 sheet writes DIVRE parents in DAOP notation. `VV`-style codes are
# what the rest of the system uses (see _ROMAN_TO_DIVRE in main.py), and "KAC"
# is Balaiyasa Kiaracondong.
PARENT_ALIAS = {
    "DI": "VI",
    "DII": "VII",
    "DIII": "VIII",
    "DIV": "VIV",
    "KAC": "BY3",
}

# `ID Lokasi` codes the workbook uses that are not `lokasi` rows. "BYK" is
# Balaiyasa Kiaracondong, whose workshop row is BY3A.
LOKASI_ALIAS = {
    "BYK": "BY3A",
}

PERUNTUKAN_MAP = {
    "A": "JALAN REL",
    "B": "JEMBATAN",
    "C": "MEKANIK",
    "D": "BALAIYASA",
}
PERUNTUKAN_KE_KODE = {v: k for k, v in PERUNTUKAN_MAP.items()}

# 160 rows in the 2026 sheet carry "PROSES" instead of a purchase date:
# procurement is still running. `aset.tanggal_pembelian` is NOT NULL and its
# year is baked into the asset ID, so there is no way to represent "unknown".
#
# These are real, planned assets, so they are imported against 1 January of the
# sheet's own year rather than dropped, and the opening RiwayatKondisi says so
# in plain Indonesian, so nobody later reads the date as a fact.
PROSES_FALLBACK = date(2026, 1, 1)
PROSES_NOTE = "Aset Baru — tanggal pembelian belum final (PROSES saat impor)."


def _clean_code(value) -> str:
    """Strip every internal space and upper-case. 'JB 1.4' → 'JB1.4'."""
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).upper()


def _clean_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace(" ", " ")).strip()


def parse_tanggal(value):
    """
    → date, or None when the cell holds no date at all.

    Ambiguity note: "21/4/26" and "24/11/2025" are both DAY-first, which is the
    Indonesian convention and consistent with every unambiguous value in the
    workbook (e.g. "17/12/2025", where 17 cannot be a month). Parsing these
    month-first would silently move assets between years.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text or not any(ch.isdigit() for ch in text):
        return None  # "PROSES" and friends

    # ISO first, it is unambiguous.
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if m:
        y, mo, d = (int(g) for g in m.groups())
        try:
            return date(y, mo, d)
        except ValueError:
            return None

    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})$", text)
    if m:
        d, mo, y = (int(g) for g in m.groups())
        if y < 100:
            # Two-digit years in this workbook are all 20xx; the fleet has no
            # 19xx purchases. Same cutoff the frontend's decodeAsetId uses.
            y += 2000 if y <= 30 else 1900
        try:
            return date(y, mo, d)
        except ValueError:
            return None
    return None


def _sheet_year(sheet_name: str):
    m = re.search(r"(20\d{2})", sheet_name)
    return int(m.group(1)) if m else None


def read_rows(path: str = WORKBOOK):
    """
    Yield one cleaned dict per data row, plus a dict of skip counters.

    Reading is separated from writing so the workbook can be validated without
    a database, `python -m seed_aset_real` below does exactly that.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    skipped = {"kode_tidak_dikenal": {}, "tanpa_parent": 0, "tanpa_kode": 0}
    out = []

    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        fallback_year = _sheet_year(sheet_name)
        for raw in list(wb[sheet_name].iter_rows(values_only=True))[1:]:
            if len(raw) <= COL_MODEL:
                continue
            kode = _clean_code(raw[COL_KODE])
            if not kode:
                continue  # padding row, not a skip worth reporting

            if kode not in ALL_KODE:
                skipped["kode_tidak_dikenal"][kode] = (
                    skipped["kode_tidak_dikenal"].get(kode, 0) + 1
                )
                continue

            parent = _clean_code(raw[COL_PARENT])
            parent = PARENT_ALIAS.get(parent, parent)
            if not parent:
                skipped["tanpa_parent"] += 1
                continue

            # Two sheets fill ID Lokasi, two do not. With no resort the asset is
            # homed at the DAOP/DIVRE itself, which the rest of the system
            # already handles (get_public_aset prints no UPT row for it).
            id_lokasi = _clean_code(raw[COL_IDLOK]) or parent
            id_lokasi = LOKASI_ALIAS.get(id_lokasi, id_lokasi)

            tanggal = parse_tanggal(raw[COL_TANGGAL])
            provisional = tanggal is None
            if provisional:
                tanggal = (
                    date(fallback_year, 1, 1) if fallback_year else PROSES_FALLBACK
                )

            # ── Peruntukan ──
            #
            # THE UPT CODE PREFIX IS THE PERUNTUKAN, and it outranks the
            # workbook's own `Unit Peruntukan` column, which is not reliable:
            # every one of the 245 rows on `Data Aset JB` carries "A" even
            # though they are bridge assets sitting on JB resorts. Importing
            # that column verbatim mints 245 permanently wrong composite
            # primary keys, peruntukan is a segment of id_aset, so it cannot
            # be corrected later without rewriting every child row.
            #
            # Verified 1:1 against the katalog's own UNIT column on all 254
            # resorts: JR -> JALAN REL, JB -> JEMBATAN, ME -> MEKANIK.
            #
            # BY* is NOT a unit. Balaiyasa is a workshop; an asset parked at
            # BY3A is visiting for repair and keeps the peruntukan it arrived
            # with. Deriving there would contradict the rule enforced in five
            # other places that a Balaiyasa is never a reporting region, so
            # for BY*, and for a bare parent code, the sheet's column wins.
            unit = _clean_code(raw[COL_UNIT])[:1] or "A"
            if unit not in PERUNTUKAN_MAP:
                unit = "A"

            # Priority: UPT code prefix, then the sheet's own identity, then the
            # Unit column. Each step is more specific than the one after it.
            turunan = peruntukan_dari_lokasi(id_lokasi) or SHEET_PERUNTUKAN.get(
                sheet_name
            )
            if turunan:
                unit = PERUNTUKAN_KE_KODE[turunan]

            pengadaan_raw = _clean_code(raw[COL_PENGADAAN])
            # "DAOP / DIVRE" survives _clean_code as "DAOP/DIVRE".
            id_pengadaan = 1 if pengadaan_raw == "PUSAT" else 2

            urut = raw[COL_URUT]
            try:
                urut = int(float(urut)) if urut is not None else None
            except (TypeError, ValueError):
                urut = None

            out.append(
                {
                    "sheet": sheet_name,
                    "kode_alat": kode,
                    "urut": urut,
                    "id_pengadaan": id_pengadaan,
                    "sumber_pengadaan": "PUSAT" if id_pengadaan == 1 else "DAOP/DIVRE",
                    "tanggal": tanggal,
                    "tanggal_provisional": provisional,
                    "unit": unit,
                    "peruntukan": PERUNTUKAN_MAP[unit],
                    "id_lokasi": id_lokasi,
                    "parent": parent,
                    "model": _clean_text(raw[COL_MODEL]),
                }
            )

    wb.close()
    return out, skipped


# ---------------------------------------------------------------------------
# Model resolution
# ---------------------------------------------------------------------------
def _model_key(text: str) -> str:
    """Loose comparison key, case, spaces and punctuation all vary in the sheet
    ('HONDA GX 160' vs 'HONDA GX160', 'Model INR - 10' vs 'INR-10')."""
    return re.sub(r"[^A-Z0-9]", "", (text or "").upper())


def resolve_model(db, kode_alat: str, model_text: str, cache: dict):
    """
    Find or create the `alat_varian` this row's Model column names.

    Matching is loose and tried in three passes, exact-ish key on
    `nama_varian`, then on `tipe_model`, then containment either way, because
    the workbook's model strings are free text typed by four different people.

    A miss CREATES a bare model row rather than leaving the asset unmodelled.
    That is deliberate: an asset with no model is invisible to the sparepart
    compatibility filter and to the spec card, whereas an obviously incomplete
    model row shows up in Pusat Data ▸ Model/Type as work to do.
    """
    if not model_text:
        return None

    if kode_alat not in cache:
        rows = db.query(models.AlatVarian).filter_by(kode_alat=kode_alat).all()
        cache[kode_alat] = rows

    key = _model_key(model_text)
    rows = cache[kode_alat]

    for row in rows:
        if _model_key(row.nama_varian) == key:
            return row
    for row in rows:
        if row.tipe_model and _model_key(row.tipe_model) == key:
            return row
    for row in rows:
        for candidate in (row.nama_varian, row.tipe_model):
            ck = _model_key(candidate)
            if ck and len(ck) >= 4 and (ck in key or key in ck):
                return row

    created = models.AlatVarian(
        kode_alat=kode_alat,
        nama_varian=model_text[:50],
        tipe_model=model_text[:100],
        keterangan="Dibuat otomatis saat impor aset; spesifikasi belum dilengkapi.",
    )
    db.add(created)
    db.flush()  # need id_varian before the asset row references it
    cache[kode_alat].append(created)
    return created


# The writer that used to live here (`seed_aset_real`) was superseded by
# seeds/aset.py::run(). It was dead for a whole release AND still carried the
# renumber-on-collision bug that doubled the fleet, the exact failure the
# seeds/ package was written to eliminate, plus a verbatim copy of 25 lines
# of seeds/aset.py, so every fix had to be made twice. This module now does
# what its docstring always claimed: it READS and CLEANS the workbook, and
# writes nothing.


if __name__ == "__main__":
    # Dry run: validate the workbook without touching the database.
    rows, skipped = read_rows()
    print(f"baris terbaca : {len(rows)}")
    print(f"dilewati      : {skipped}")
    print(f"provisional   : {sum(1 for r in rows if r['tanggal_provisional'])}")
    by_sheet = {}
    for r in rows:
        by_sheet[r["sheet"]] = by_sheet.get(r["sheet"], 0) + 1
    print(f"per sheet     : {by_sheet}")
    print(f"tahun         : {sorted({r['tanggal'].year for r in rows})}")
    print(f"parent        : {sorted({r['parent'] for r in rows})}")
